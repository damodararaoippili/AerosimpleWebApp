import openpyxl
from BaseClass import BasePage
class LoginTestData(BasePage):
    workbook_path = "C:\\Users\\damod\\Aerosimple Web Application\\TestData\\AerosimpleTestData.xlsx"
    worksheet_name = "LoginDetails"

    def __init__(self, driver):
        super().__init__(driver)

    def read_excel_data(self):
        workbook = openpyxl.load_workbook(LoginTestData.workbook_path)
        worksheet = workbook[LoginTestData.worksheet_name]
        rows = worksheet.max_row
        columns = worksheet.max_column
        dict_data = {}

        for row in range(1, rows + 1):
            row_data = {}
            for col in range(1, columns + 1):
                key = worksheet.cell(row=row, column=col).value
                value = worksheet.cell(row=row, column=col + 1).value
                row_data[key] = value
            dict_data.update(row_data)

        return dict_data
